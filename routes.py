from flask import Blueprint, request, render_template, redirect, url_for, flash, session, Response, current_app, send_from_directory
from models import db, Patient, EstimationEntry
from dental_methods import calculate_demirjian_score, calculate_alqahtani_age, get_alqahtani_teeth, get_demirjian_teeth
from functools import wraps
from sqlalchemy import cast
import random
import string
import csv
from io import StringIO, BytesIO
import os
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
# Add openpyxl imports for Excel generation and reading
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
# Add matplotlib for chart generation
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import numpy as np
import logging
import datetime

main = Blueprint('main', __name__)

# Configure logging
try:
    logger = current_app.logger
except:
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

# Simple in-memory cache for charts (in production, use Redis or similar)
chart_cache = {}

# Cache timeout in seconds (e.g., 1 hour)
CHART_CACHE_TIMEOUT = 3600

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('auth.login'))
        return func(*args, **kwargs)
    return wrapper

def role_required(role):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('auth.login'))
            if session.get('role') != role and session.get('role') != 'supervisor':
                flash('Access denied')
                return redirect(url_for('main.dashboard'))
            return func(*args, **kwargs)
        return wrapper
    return decorator

def generate_csrf_token():
    """Generate and store CSRF token in session"""
    if 'csrf_token' not in session:
        session['csrf_token'] = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
    return session['csrf_token']

@main.context_processor
def inject_csrf_token():
    """Inject CSRF token into all templates"""
    return dict(csrf_token=generate_csrf_token())

def validate_csrf_token():
    """Validate CSRF token for POST requests"""
    token = session.get('csrf_token')
    
    # Handle different content types
    if request.is_json:
        # For JSON requests, get token from headers or JSON body
        form_token = request.headers.get('X-CSRF-Token') or request.get_json(silent=True).get('csrf_token') if request.get_json(silent=True) else None
    else:
        # For form submissions, get token from form data or headers
        form_token = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
    
    if not token or not form_token or token != form_token:
        return False
    return True

@main.before_request
def csrf_protect():
    """CSRF protection for all POST requests"""
    # TEMPORARILY DISABLED FOR DEBUGGING
    return None
    # if request.method == "POST":
    #     # Skip CSRF check for specific endpoints
    #     if request.endpoint not in ['main.upload_opg', 'main.get_estimation_form']:
    #         if not validate_csrf_token():
    #             # Determine response format based on request type
    #             if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
    #                 # Return JSON error for AJAX or JSON requests
    #                 return Response(
    #                     '{"error": "Security token validation failed. Please refresh the page and try again."}', 
    #                     status=400, 
    #                     mimetype='application/json'
    #                 )
    #             # For regular form submissions, use flash and redirect
    #             flash('Security token validation failed. Please try again.')
    #             return redirect(request.url)

@main.route('/dashboard')
@login_required
def dashboard():
    if session['role'] == 'supervisor':
        return render_template('supervisor_dashboard.html')
    else:
        return render_template('pi_dashboard.html')

def renumber_patient_ids():
    """Renumber patient IDs sequentially starting from 1"""
    patients = Patient.query.order_by(Patient.id).all()
    
    # First, assign temporary IDs to avoid unique constraint violations
    for i, patient in enumerate(patients, 1):
        # Assign a temporary ID that won't conflict
        patient.patient_id = f"TEMP_{patient.id}"
    db.session.flush()  # Flush to database but don't commit yet
    
    # Then assign the final sequential IDs
    for i, patient in enumerate(patients, 1):
        # Check if patient_id is in format "T<number>" or just a number
        if patient.patient_id.startswith('T') and not patient.patient_id.startswith('TEMP_'):
            # Update T-prefixed IDs to be sequential
            patient.patient_id = f"T{i}"
        elif patient.patient_id.startswith('TEMP_'):
            # This is a patient that had a numeric ID before
            # Update numeric IDs to be sequential
            patient.patient_id = str(i)
        else:
            # Check if it's a pure number
            try:
                # Try to convert to int to check if it's a pure number
                int(patient.patient_id)
                patient.patient_id = str(i)
            except ValueError:
                # If it's not a pure number, keep as is but log
                pass
    
    db.session.commit()

@main.route('/patients', methods=['GET', 'POST'])
@role_required('supervisor')
def manage_patients():
    if request.method == 'POST':
        # Handle file upload
        if 'csv_file' in request.files and request.files['csv_file'].filename != '':
            file = request.files['csv_file']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                if filename.endswith(('.csv',)):
                    # Process CSV upload
                    stream = StringIO(file.stream.read().decode("UTF8"), newline=None)
                    csv_input = csv.reader(stream)
                    
                    # Skip header row
                    next(csv_input, None)
                    
                    added_count = 0
                    skipped_count = 0
                    
                    for row in csv_input:
                        # Handle different CSV formats:
                        # Format 1 (Simple): ID, Name, Actual Age, Sex (4 columns)
                        # Format 2 (Full): ID, Name, Age, Sex, OPG, A code, D code, A Age, D Age, Actual age (10+ columns)
                        
                        patient_id = ''
                        name = ''
                        actual_age = ''
                        sex = ''
                        code_a = ''
                        code_b = ''
                        
                        if len(row) >= 4:  # Simple format
                            patient_id = row[0]
                            name = row[1]
                            actual_age = row[2]
                            sex = row[3]
                            
                            # Check if it's the full format (10+ columns)
                            if len(row) >= 10:
                                # Override with full format data
                                patient_id = row[0]
                                name = row[1]
                                actual_age = row[9]
                                sex = row[3]
                                code_a = row[5]
                                code_b = row[6]
                            
                            # Check if patient already exists
                            patient = Patient.query.filter_by(patient_id=patient_id).first()
                            if not patient:
                                patient = Patient(
                                    patient_id=patient_id,
                                    name=name,
                                    actual_age=float(actual_age) if actual_age else 0,
                                    sex=sex,
                                    code_a=code_a if code_a else None,
                                    code_b=code_b if code_b else None
                                )
                                db.session.add(patient)
                                added_count += 1
                                # Commit in batches to handle large datasets
                                if added_count % 100 == 0:
                                    db.session.flush()
                            else:
                                skipped_count += 1
                    
                    try:
                        db.session.commit()
                        flash(f'CSV import successful! Added: {added_count}, Skipped (already exist): {skipped_count}')
                    except Exception as e:
                        db.session.rollback()
                        flash(f'Error importing CSV: {str(e)}')
                        
                elif filename.endswith(('.xlsx', '.xls')):
                    # Process Excel upload
                    try:
                        # Save file to a temporary location to process with openpyxl
                        file.stream.seek(0)
                        temp_file_path = os.path.join('/tmp', filename)
                        file.save(temp_file_path)
                        
                        # Load workbook
                        workbook = openpyxl.load_workbook(temp_file_path, data_only=True)
                        worksheet = workbook.active
                        
                        # Skip header row
                        first_row = True
                        row_count = 0
                        added_count = 0
                        skipped_count = 0
                        
                        for row in worksheet.iter_rows(values_only=True):
                            if first_row:
                                first_row = False
                                continue
                            
                            row_count += 1
                            # Handle different Excel formats:
                            # Format 1 (Simple): ID, Name, Actual Age, Sex (4 columns)
                            # Format 2 (Full): ID, Name, Age, Sex, OPG, A code, D code, A Age, D Age, Actual age (10+ columns)
                            
                            patient_id = ''
                            name = ''
                            actual_age = ''
                            sex = ''
                            code_a = ''
                            code_b = ''
                            
                            if len(row) >= 4:  # Simple format
                                patient_id = str(row[0]) if row[0] else ''
                                name = str(row[1]) if row[1] else ''
                                actual_age = str(row[2]) if row[2] else ''
                                sex = str(row[3]) if row[3] else ''
                                
                                # Check if it's the full format (10+ columns)
                                if len(row) >= 10:
                                    # Override with full format data
                                    patient_id = str(row[0]) if row[0] else ''
                                    name = str(row[1]) if row[1] else ''
                                    actual_age = str(row[9]) if row[9] else ''  # Column 9 is actual age
                                    sex = str(row[3]) if row[3] else ''  # Column 3 is sex
                                    code_a = str(row[5]) if row[5] else ''  # Column 5 is A code
                                    code_b = str(row[6]) if row[6] else ''  # Column 6 is D code
                                
                                # Check if we have a patient_id
                                if patient_id:
                                    # Check if patient already exists
                                    patient = Patient.query.filter_by(patient_id=patient_id).first()
                                    if not patient:
                                        patient = Patient(
                                            patient_id=patient_id,
                                            name=name,
                                            actual_age=float(actual_age) if actual_age and actual_age.replace('.', '', 1).isdigit() else 0,
                                            sex=sex,
                                            code_a=code_a if code_a else None,  # Ensure None if empty
                                            code_b=code_b if code_b else None   # Ensure None if empty
                                        )
                                        db.session.add(patient)
                                        added_count += 1
                                        # Commit in batches to handle large datasets
                                        if added_count % 100 == 0:
                                            db.session.flush()
                                    else:
                                        skipped_count += 1
                        
                        workbook.close()
                        # Clean up temp file
                        os.remove(temp_file_path)
                        
                        try:
                            db.session.commit()
                            flash(f'Excel import successful! Added: {added_count}, Skipped (already exist): {skipped_count}')
                        except Exception as e:
                            db.session.rollback()
                            flash(f'Error importing Excel: {str(e)}')
                    except Exception as e:
                        # Clean up temp file on error
                        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                            os.remove(temp_file_path)
                        flash(f'Error processing Excel file: {str(e)}')
                        
        else:
            # Handle manual patient creation
                
            patient_id = request.form['patient_id']
            name = request.form.get('name', '')
            actual_age = float(request.form['actual_age'])
            sex = request.form['sex']
            
            # Check if patient already exists
            patient = Patient.query.filter_by(patient_id=patient_id).first()
            if not patient:
                patient = Patient(
                    patient_id=patient_id,
                    name=name,
                    actual_age=actual_age,
                    sex=sex
                )
                db.session.add(patient)
                try:
                    db.session.commit()
                    flash('Patient added successfully!')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding patient: {str(e)}')
            else:
                flash('Patient with this ID already exists!')
        
        # Renumber patient IDs after any addition
        renumber_patient_ids()
        
        return redirect(url_for('main.manage_patients'))
    
    # Handle search functionality and pagination
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Show 20 entries per page
    
    # Query patients with search and pagination
    patients_query = Patient.query
    
    if search_query:
        patients_query = patients_query.filter(
            db.or_(
                Patient.patient_id.contains(search_query),
                Patient.name.contains(search_query),
                Patient.code_a.contains(search_query),
                Patient.code_b.contains(search_query)
            )
        )
    
    # Order patients by patient_id numerically
    patients_query = patients_query.order_by(cast(Patient.patient_id, db.Integer))
    
    # Get patients with pagination
    patients = patients_query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the content part for AJAX requests
        return render_template('patients_content.html', patients=patients, search_query=search_query)
    
    return render_template('patients.html', patients=patients, search_query=search_query)

@main.route('/patients/update/<patient_id>', methods=['GET', 'POST'])
@role_required('supervisor')
def update_patient(patient_id):
    current_app.logger.info(f"Entered update_patient with id: {patient_id}")
    try:
        patient_id_int = int(patient_id)
        patient = Patient.query.get_or_404(patient_id_int)
    except ValueError:
        current_app.logger.error(f"Invalid patient_id for update: {patient_id}")
        return f"Invalid Patient ID: {patient_id}", 400
    
    if request.method == 'POST':
        old_patient_id = patient.patient_id
        
        # Update patient details
        patient.patient_id = request.form['patient_id']
        patient.name = request.form.get('name', '')
        patient.actual_age = float(request.form['actual_age'])
        patient.sex = request.form['sex']
        
        try:
            db.session.commit()
            flash('Patient updated successfully!')
            
            # If patient_id was changed, renumber all patient IDs
            if old_patient_id != patient.patient_id:
                renumber_patient_ids()
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating patient: {str(e)}')
        
        return redirect(url_for('main.manage_patients'))
    
    return render_template('update_patient.html', patient=patient)

@main.route('/patients/delete/<patient_id>', methods=['POST'])
@role_required('supervisor')
def delete_patient(patient_id):
    current_app.logger.info(f"Entered delete_patient with id: {patient_id}")
    try:
        patient_id_int = int(patient_id)
        patient = Patient.query.get_or_404(patient_id_int)
    except ValueError:
        current_app.logger.error(f"Invalid patient_id for delete: {patient_id}")
        return f"Invalid Patient ID: {patient_id}", 400
    
    # Delete OPG image from Supabase if it exists
    if patient.opg_link and patient.opg_link.startswith('http'):
        try:
            from utils.storage import delete_image
            # Extract filename from URL - works for both public URLs and signed URLs
            # For signed URLs, we need to extract the path before the query parameters
            if '?' in patient.opg_link:
                # Signed URL - extract filename from path before query parameters
                path_part = patient.opg_link.split('?')[0]
                filename = path_part.split('/')[-1]
            else:
                # Regular URL - extract filename from last part of URL
                filename = patient.opg_link.split('/')[-1]
            delete_image(filename)
        except:
            pass  # If deletion fails, continue anyway
    
    # Delete OPG image file if it exists locally (for backward compatibility)
    if patient.opg_link and not patient.opg_link.startswith('http'):
        try:
            os.remove(os.path.join(current_app.root_path, patient.opg_link))
        except:
            pass  # If file doesn't exist or can't be deleted, continue anyway
    
    # Delete associated estimation entries (orphans)
    if patient.code_a:
        EstimationEntry.query.filter_by(code=patient.code_a).delete()
    if patient.code_b:
        EstimationEntry.query.filter_by(code=patient.code_b).delete()
        
    # Delete the patient record
    db.session.delete(patient)
    
    try:
        db.session.commit()
        
        # Clear chart cache so charts update instantly
        global chart_cache
        chart_cache = {}
        
        flash('Patient and associated data deleted successfully!')
        
        # Renumber patient IDs after deletion
        renumber_patient_ids()
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting patient: {str(e)}')
    
    return redirect(url_for('main.manage_patients'))

@main.route('/assign_codes')
@role_required('supervisor')
def assign_codes():
    patients = Patient.query.filter((Patient.code_a.is_(None)) | (Patient.code_b.is_(None))).all()
    
    for patient in patients:
        if not patient.code_a:
            # Generate random code for AlQahtani method
            patient.code_a = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        
        if not patient.code_b:
            # Generate random code for Demirjian method
            patient.code_b = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    
    db.session.commit()
    flash('Codes assigned to all patients')
    return redirect(url_for('main.manage_patients'))

@main.route('/upload_opg/<patient_id>', methods=['GET', 'POST'])
@role_required('supervisor')
def upload_opg(patient_id):
    current_app.logger.info(f"Entered upload_opg for patient {patient_id}")
    try:
        patient_id_int = int(patient_id)
        patient = Patient.query.get_or_404(patient_id_int)
    except ValueError:
        current_app.logger.error(f"Invalid patient_id for upload: {patient_id}")
        return f"Invalid Patient ID: {patient_id}", 400
    
    if request.method == 'POST':
        current_app.logger.info("Processing POST request in upload_opg")
        # Handle file upload
        if 'opg_file' in request.files:
            file = request.files['opg_file']
            
            # Validate file
            if not file:
                current_app.logger.error("No file object received")
                flash('No file selected')
                return redirect(url_for('main.upload_opg', patient_id=patient_id))
            
            if file.filename == '':
                current_app.logger.error("Empty filename received")
                flash('No file selected')
                return redirect(url_for('main.upload_opg', patient_id=patient_id))
            
            if not allowed_file(file.filename):
                current_app.logger.error(f"Invalid file type: {file.filename}")
                flash('Invalid file type. Please upload an image file.')
                return redirect(url_for('main.upload_opg', patient_id=patient_id))
            
            current_app.logger.info(f"Processing OPG upload for patient {patient_id}, file: {file.filename}")
            
            filename = secure_filename(f"{patient.patient_id}_{file.filename}")
            current_app.logger.info(f"Secured filename: {filename}")
            
            try:
                # Import Supabase storage utility
                from utils.storage import upload_image, delete_image
                current_app.logger.info("Imported Supabase storage utilities")
                
                # If patient already has an OPG image, delete it from Supabase first
                if patient.opg_link and patient.opg_link.startswith('http'):
                    try:
                        current_app.logger.info(f"Patient has existing OPG link: {patient.opg_link[:50]}...")
                        # Extract filename from URL - works for both public URLs and signed URLs
                        # For signed URLs, we need to extract the path before the query parameters
                        if '?' in patient.opg_link:
                            # Signed URL - extract filename from path before query parameters
                            path_part = patient.opg_link.split('?')[0]
                            old_filename = path_part.split('/')[-1]
                        else:
                            # Regular URL - extract filename from last part of URL
                            old_filename = patient.opg_link.split('/')[-1]
                        current_app.logger.info(f"Attempting to delete old image: {old_filename}")
                        delete_image(old_filename)
                        current_app.logger.info("Old image deleted successfully")
                    except Exception as e:
                        # Log error but continue with upload
                        current_app.logger.error(f"Failed to delete old image: {str(e)}")
                        current_app.logger.error(f"Error type: {type(e)}")
                
                # Upload to Supabase (no local storage needed on Vercel)
                current_app.logger.info(f"Starting upload to Supabase for file: {filename}")
                file_url = upload_image(file, filename)
                current_app.logger.info(f"Upload successful! File URL: {file_url[:100]}...")
                
                # Store URL in database
                patient.opg_link = file_url
                db.session.commit()
                current_app.logger.info(f"Database updated with new OPG link for patient {patient_id}")
                
                flash('OPG uploaded successfully')
                return redirect(url_for('main.manage_patients'))
                
            except ValueError as ve:
                # Supabase configuration error
                current_app.logger.error(f"Supabase configuration error: {str(ve)}")
                flash(f'Configuration error: {str(ve)}. Please contact administrator.')
                return redirect(url_for('main.upload_opg', patient_id=patient_id))
            except Exception as e:
                current_app.logger.error(f"Error uploading OPG: {str(e)}")
                current_app.logger.error(f"Error type: {type(e)}")
                import traceback
                current_app.logger.error(f"Traceback: {traceback.format_exc()}")
                flash(f'Error uploading OPG: {str(e)}')
                return redirect(url_for('main.upload_opg', patient_id=patient_id))
    
    return render_template('upload_opg.html', patient=patient)

# Serve uploaded files
@main.route('/uploads/<filename>')
def uploaded_file(filename):
    # Only allow access to authenticated users
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # Return the file from the uploads directory
    # For backward compatibility with local files
    try:
        return send_from_directory(UPLOAD_FOLDER, filename)
    except:
        flash('File not found')
        return redirect(url_for('main.manage_patients'))

@main.route('/blinded_data')
@role_required('supervisor')
def blinded_data():
    # Handle search functionality and pagination
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Show 20 entries per page
    
    # Generate blinded data for PI
    patients_query = Patient.query.filter(
        db.or_(
            Patient.code_a.isnot(None), 
            Patient.code_b.isnot(None)
        )
    )
    
    if search_query:
        patients_query = patients_query.filter(
            db.or_(
                Patient.code_a.contains(search_query),
                Patient.code_b.contains(search_query)
            )
        )
    
    # Get patients with pagination
    patients = patients_query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    blinded_entries = []
    for patient in patients.items:
        # Add AlQahtani entry if code exists
        if patient.code_a:
            entry_a = {
                'code': patient.code_a,
                'opg_link': patient.opg_link,
                'sex': patient.sex,
                'method': 'AlQahtani'
            }
            blinded_entries.append(entry_a)
        
        # Add Demirjian entry if code exists
        if patient.code_b:
            entry_b = {
                'code': patient.code_b,
                'opg_link': patient.opg_link,
                'sex': patient.sex,
                'method': 'Demirjian'
            }
            blinded_entries.append(entry_b)
    
    # Shuffle the data
    random.shuffle(blinded_entries)
    
    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the content part for AJAX requests
        return render_template('blinded_data_content.html', entries=blinded_entries, search_query=search_query, patients=patients)
    
    return render_template('blinded_data.html', entries=blinded_entries, search_query=search_query, patients=patients)

@main.route('/estimate_age', methods=['GET', 'POST'])
@role_required('pi')
def estimate_age():
    if request.method == 'POST':
        # Validate CSRF token
        if not validate_csrf_token():
            flash('Security token validation failed. Please try again.')
            return redirect(url_for('main.estimate_age'))
            
        code = request.form['code']
        estimated_age = float(request.form['estimated_age'])
        method = request.form['method']
        
        # Create estimation entry
        estimation = EstimationEntry(
            code=code,
            estimated_age=estimated_age,
            method_used=method
        )
        
        # Note: We're no longer collecting individual tooth stage data
        # The PI only provides the final estimated age
        
        db.session.add(estimation)
        db.session.commit()
        
        # Update patient record with estimated age
        patient = Patient.query.filter(
            (Patient.code_a == code) | (Patient.code_b == code)
        ).first()
        
        if patient:
            if code == patient.code_a:
                patient.alqahtani_estimated_age = estimated_age
            elif code == patient.code_b:
                patient.demirjian_estimated_age = estimated_age
        
        db.session.commit()
        
        flash('Estimation submitted successfully')
        return redirect(url_for('main.estimate_age'))
    
    # Handle search functionality and pagination
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Show 20 entries per page
    
    # Show blinded data for the PI to work with
    # Get all patients that have codes assigned but not yet estimated
    patients_query = Patient.query.filter(
        Patient.code_a.isnot(None), 
        Patient.code_b.isnot(None)
    )
    
    if search_query:
        patients_query = patients_query.filter(
            db.or_(
                Patient.code_a.contains(search_query),
                Patient.code_b.contains(search_query)
            )
        )
    
    # Apply pagination
    patients = patients_query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    # Find which codes have already been estimated
    existing_estimations = EstimationEntry.query.with_entities(EstimationEntry.code).all()
    estimated_codes = [e.code for e in existing_estimations]
    
    # Prepare blinded entries that haven't been estimated yet
    blinded_entries = []
    for patient in patients.items:
        # Add AlQahtani entry if code exists and not estimated yet
        if patient.code_a and patient.code_a not in estimated_codes:
            entry = {
                'code': patient.code_a,
                'opg_link': patient.opg_link,
                'sex': patient.sex,
                'method': 'AlQahtani'
            }
            blinded_entries.append(entry)
        
        # Add Demirjian entry if code exists and not estimated yet
        if patient.code_b and patient.code_b not in estimated_codes:
            entry = {
                'code': patient.code_b,
                'opg_link': patient.opg_link,
                'sex': patient.sex,
                'method': 'Demirjian'
            }
            blinded_entries.append(entry)
    
    # Shuffle the data
    random.shuffle(blinded_entries)
    
    # Remove the artificial limit that was causing issues
    # The pagination will handle the appropriate number of entries to display
    
    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the content part for AJAX requests
        return render_template('estimate_age_content.html', entries=blinded_entries, search_query=search_query, patients=patients)
    
    return render_template('estimate_age.html', entries=blinded_entries, search_query=search_query, patients=patients)

@main.route('/perform_estimation')
@role_required('pi')
def perform_estimation():
    code = request.args.get('code')
    method = request.args.get('method')
    opg = request.args.get('opg')
    sex = request.args.get('sex')
    
    # Get teeth based on method
    if method == 'AlQahtani':
        teeth = get_alqahtani_teeth()
    else:  # Demirjian
        teeth = get_demirjian_teeth()
    
    return render_template('perform_estimation.html', 
                          code=code, 
                          method=method, 
                          opg=opg, 
                          sex=sex, 
                          teeth=teeth)

def generate_chart_data():
    """Generate chart data with caching"""
    cache_key = "analysis_charts"
    
    # Check if we have cached data
    if cache_key in chart_cache:
        cached_data, timestamp = chart_cache[cache_key]
        # Cache for 5 minutes
        if (datetime.datetime.utcnow() - timestamp).total_seconds() < 300:
            logger.info("Returning cached chart data")
            return cached_data
    
    # Generate new chart data
    logger.info("Generating new chart data")
    
    # Get all estimation entries
    entries = EstimationEntry.query.all()
    
    if not entries:
        return None, None, None, None
    
    # Prepare data for charts
    alq_ages = [e.estimated_age for e in entries if e.method_used.lower() == 'alqahtani']
    dem_ages = [e.estimated_age for e in entries if e.method_used.lower() == 'demirjian']
    
    # Actual ages from patients
    actual_ages = []
    for e in entries:
        patient = Patient.query.filter(
            (Patient.code_a == e.code) | (Patient.code_b == e.code)
        ).first()
        if patient:
            actual_ages.append(patient.actual_age)
    
    # Generate charts
    # 1. Age Distribution Chart
    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    
    if alq_ages or dem_ages:
        if alq_ages:
            ax.hist(alq_ages, bins=20, alpha=0.7, label='AlQahtani', color='#2563eb')
        if dem_ages:
            ax.hist(dem_ages, bins=20, alpha=0.7, label='Demirjian', color='#818cf8')
        ax.set_xlabel('Estimated Age (years)')
        ax.set_ylabel('Frequency')
        ax.set_title('Distribution of Estimated Ages')
        ax.legend()
    
    buf = BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    age_dist_chart = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    
    # 2. Actual vs Estimated Chart
    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    
    if actual_ages and (alq_ages or dem_ages):
        if alq_ages and len(actual_ages) == len(alq_ages):
            ax.scatter(actual_ages[:len(alq_ages)], alq_ages, alpha=0.7, label='AlQahtani', color='#2563eb')
        if dem_ages and len(actual_ages) == len(dem_ages):
            ax.scatter(actual_ages[:len(dem_ages)], dem_ages, alpha=0.7, label='Demirjian', color='#818cf8')
        
        # Perfect prediction line
        if actual_ages:
            min_age = min(actual_ages)
            max_age = max(actual_ages)
            ax.plot([min_age, max_age], [min_age, max_age], 'r--', label='Perfect Prediction')
        
        ax.set_xlabel('Actual Age (years)')
        ax.set_ylabel('Estimated Age (years)')
        ax.set_title('Actual vs Estimated Ages')
        ax.legend()
    
    buf = BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    actual_vs_estimated_chart = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    
    # 3. Error Distribution Chart
    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    
    alq_errors = []
    dem_errors = []
    
    for i, e in enumerate(entries):
        patient = Patient.query.filter(
            (Patient.code_a == e.code) | (Patient.code_b == e.code)
        ).first()
        if patient:
            error = abs(e.estimated_age - patient.actual_age)
            if e.method_used.lower() == 'alqahtani':
                alq_errors.append(error)
            elif e.method_used.lower() == 'demirjian':
                dem_errors.append(error)
    
    if alq_errors or dem_errors:
        if alq_errors:
            ax.hist(alq_errors, bins=20, alpha=0.7, label='AlQahtani', color='#2563eb')
        if dem_errors:
            ax.hist(dem_errors, bins=20, alpha=0.7, label='Demirjian', color='#818cf8')
        ax.set_xlabel('Absolute Error (years)')
        ax.set_ylabel('Frequency')
        ax.set_title('Distribution of Estimation Errors')
        ax.legend()
    
    buf = BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    error_dist_chart = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    
    # 4. Method Comparison Chart
    plt.clf()
    fig, ax = plt.subplots(figsize=(10, 6))
    
    alq_mean_error = np.mean(alq_errors) if alq_errors else 0
    dem_mean_error = np.mean(dem_errors) if dem_errors else 0
    
    methods = ['AlQahtani', 'Demirjian']
    mean_errors = [alq_mean_error, dem_mean_error]
    colors = ['#2563eb', '#818cf8']
    
    bars = ax.bar(methods, mean_errors, color=colors)
    ax.set_ylabel('Mean Absolute Error (years)')
    ax.set_title('Comparison of Methods')
    
    # Add value labels on bars
    for bar, error in zip(bars, mean_errors):
        height = bar.get_height()
        ax.annotate(f'{error:.2f}',
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha='center', va='bottom')
    
    buf = BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    method_comparison_chart = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    
    chart_data = (age_dist_chart, actual_vs_estimated_chart, error_dist_chart, method_comparison_chart)
    
    # Cache the data
    chart_cache[cache_key] = (chart_data, datetime.datetime.utcnow())
    
    return chart_data

@main.route('/analysis')
@role_required('supervisor')
def analysis():
    # Handle search functionality and pagination
    search_query = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Show 20 entries per page
    
    # Query patients with search and pagination
    patients_query = Patient.query
    
    if search_query:
        patients_query = patients_query.filter(
            db.or_(
                Patient.patient_id.contains(search_query),
                Patient.name.contains(search_query),
                Patient.code_a.contains(search_query),
                Patient.code_b.contains(search_query)
            )
        )
    
    # Order patients by patient_id numerically
    patients_query = patients_query.order_by(cast(Patient.patient_id, db.Integer))
    
    # Get patients with pagination
    patients = patients_query.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    # Prepare results for the template
    results = patients.items
    
    # Generate or retrieve cached chart data
    chart_data = generate_chart_data()
    
    if chart_data[0] is None:
        flash('No data available for analysis yet.')
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Return only the content part for AJAX requests
            return render_template('analysis_content.html',
                                 accuracy_chart=None,
                                 comparison_chart=None,
                                 patients=patients,
                                 results=results,
                                 search_query=search_query)
        
        return render_template('analysis.html', 
                             accuracy_chart=None,
                             comparison_chart=None,
                             patients=patients,
                             results=results,
                             search_query=search_query)
    
    age_dist_chart, actual_vs_estimated_chart, error_dist_chart, method_comparison_chart = chart_data
    
    # Map the generated charts to the expected template variables
    accuracy_chart = age_dist_chart  # Using age distribution as the accuracy chart
    comparison_chart = method_comparison_chart  # Using method comparison as the comparison chart
    
    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the content part for AJAX requests
        return render_template('analysis_content.html',
                             accuracy_chart=accuracy_chart,
                             comparison_chart=comparison_chart,
                             patients=patients,
                             results=results,
                             search_query=search_query)
    
    return render_template('analysis.html',
                         accuracy_chart=accuracy_chart,
                         comparison_chart=comparison_chart,
                         patients=patients,
                         results=results,
                         search_query=search_query)

@main.route('/clear_chart_cache')
@role_required('supervisor')
def clear_chart_cache():
    """Clear the chart cache"""
    global chart_cache
    chart_cache = {}
    flash('Chart cache cleared successfully.')
    return redirect(url_for('main.analysis'))

@main.route('/export_patients')
@role_required('supervisor')
def export_patients():
    # Export all patients as Excel file with embedded images
    # Process in batches to handle large datasets
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Data"
    
    # Write header
    headers = ['ID', 'Name', 'Age', 'Sex', 'OPG Image', 'A code', 'D code', 'A Age', 'D Age', 'Actual age']
    ws.append(headers)
    
    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Increase width for Image column (5th column -> E)
    ws.column_dimensions['E'].width = 25

    # Make header row bold
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = cell.font.copy(bold=True)
    
    # Process patients in batches to handle large datasets
    offset = 0
    batch_size = 100
    has_more = True
    
    # Retry logic for database operations
    max_retries = 3
    retry_delay = 1  # seconds
    
    while has_more:
        # Order patients by patient_id numerically to ensure sequential order in export
        patients = None
        for attempt in range(max_retries):
            try:
                patients = Patient.query.order_by(cast(Patient.patient_id, db.Integer)).offset(offset).limit(batch_size).all()
                break  # Success, exit retry loop
            except Exception as e:
                if "SSL error" in str(e) and attempt < max_retries - 1:
                    # SSL error, wait and retry
                    import time
                    time.sleep(retry_delay * (2 ** attempt))  # Exponential backoff
                    continue
                else:
                    # Re-raise the exception if it's not an SSL error or we've exhausted retries
                    raise
        
        if not patients:
            has_more = False
            continue
            
        # Write data
        for row_idx, patient in enumerate(patients, start=offset+2):
            # Set row height to accommodate image
            ws.row_dimensions[row_idx].height = 80
            
            # Add patient data
            ws.cell(row=row_idx, column=1, value=patient.patient_id)
            ws.cell(row=row_idx, column=2, value=patient.name)
            ws.cell(row=row_idx, column=3, value=patient.actual_age)
            ws.cell(row=row_idx, column=4, value=patient.sex)
            # Image column (5) will be handled separately
            ws.cell(row=row_idx, column=6, value=patient.code_a)
            ws.cell(row=row_idx, column=7, value=patient.code_b)
            ws.cell(row=row_idx, column=8, value=patient.alqahtani_estimated_age)
            ws.cell(row=row_idx, column=9, value=patient.demirjian_estimated_age)
            ws.cell(row=row_idx, column=10, value=patient.actual_age)
            
            # Handle image embedding if image exists
            if patient.opg_link:
                try:
                    if patient.opg_link.startswith('http'):
                        # Handle Supabase image URL
                        try:
                            import urllib.request
                            import ssl
                            
                            # Create unverified SSL context to bypass certificate issues
                            ctx = ssl.create_default_context()
                            ctx.check_hostname = False
                            ctx.verify_mode = ssl.CERT_NONE
                            
                            # Download image data with unverified context
                            with urllib.request.urlopen(patient.opg_link, context=ctx, timeout=60) as response:
                                image_data = BytesIO(response.read())
                                # Load image
                                img = ExcelImage(image_data)
                                
                                # Resize image to fit in the cell
                                img.height = 100
                                img.width = 100
                                
                                # Add image to the worksheet at column E
                                ws.add_image(img, f'E{row_idx}')
                        except Exception as e:
                            ws.cell(row=row_idx, column=5, value="Image Load Error")
                            logger.error(f"Failed to embed image for patient {patient.patient_id}: {str(e)}")
                    else:
                        # Handle local image file
                        image_path = patient.opg_link.lstrip('/')
                        # Check multiple path possibilities
                        possible_paths = [
                            os.path.join(current_app.root_path, image_path),
                            os.path.abspath(image_path),
                            image_path
                        ]
                        
                        found_path = None
                        for path in possible_paths:
                            if os.path.exists(path):
                                found_path = path
                                break
                                
                        if found_path:
                            try:
                                img = ExcelImage(found_path)
                                img.height = 100
                                img.width = 100
                                ws.add_image(img, f'E{row_idx}')
                            except Exception as e:
                                ws.cell(row=row_idx, column=5, value="Image Format Error")
                        else:
                             ws.cell(row=row_idx, column=5, value="File Not Found")
                except Exception as e:
                    logger.error(f"Error processing image for excel: {str(e)}")
                    ws.cell(row=row_idx, column=5, value="Error")
            else:
                ws.cell(row=row_idx, column=5, value="No Image")
        
        offset += len(patients)
        if len(patients) < batch_size:
            has_more = False
    
    # Save the workbook to a BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=patients.xlsx"}
    )

@main.route('/reset-passwords', methods=['GET', 'POST'])
def reset_passwords():
    """Temporary route to reset default user passwords"""
    if request.method == 'POST':
        # Only allow this in development or with explicit environment variable
        if os.environ.get('FLASK_ENV') != 'development' and not os.environ.get('ALLOW_PASSWORD_RESET'):
            flash('Password reset not allowed in production without ALLOW_PASSWORD_RESET environment variable')
            return redirect(url_for('auth.login'))
            
        supervisor = User.query.filter_by(username='supervisor').first()
        pi = User.query.filter_by(username='pi').first()
        
        if supervisor:
            supervisor.password = generate_password_hash('supervisor')
            
        if pi:
            pi.password = generate_password_hash('pi')
            
        db.session.commit()
        flash('Passwords reset to default values. Username: supervisor, Password: supervisor for supervisor. Username: pi, Password: pi for PI.')
        return redirect(url_for('auth.login'))
    
    return '''
    <div style="padding: 20px;">
        <h2>Reset Default User Passwords</h2>
        <p>This will reset the passwords for supervisor and pi users to their default values.</p>
        <form method="POST">
            <input type="submit" value="Reset Passwords" 
                   onclick="return confirm('Are you sure you want to reset the passwords?')">
        </form>
        <p><a href="/">Back to Login</a></p>
    </div>
    '''

@main.route('/view-users')
def view_users():
    """Temporary route to view user information"""
    # Only allow this in development or with explicit environment variable
    if os.environ.get('FLASK_ENV') != 'development' and not os.environ.get('ALLOW_PASSWORD_RESET'):
        flash('User view not allowed in production without ALLOW_PASSWORD_RESET environment variable')
        return redirect(url_for('auth.login'))
    
    users = User.query.all()
    users_html = ""
    for user in users:
        users_html += f"<p>ID: {user.id}, Username: {user.username}, Role: {user.role}, Password Hash: {user.password}</p>"
    
    return f'''
    <div style="padding: 20px;">
        <h2>User Information</h2>
        {users_html}
        <p><a href="/">Back to Login</a></p>
    </div>
    '''
