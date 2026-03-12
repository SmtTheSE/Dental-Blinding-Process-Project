"""
upload_opg_and_build_excel.py
--------------------------------
Reads Blinding_TT1_mapped_links.xlsx, finds the OPG image inside each patient's
folder, uploads it to Supabase, and writes Blinding_TT1_supabase_urls.xlsx
with real https:// URLs in the OPG column.

Usage:
  SUPABASE_URL=https://hlxacefixdrvljnxqaaq.supabase.co \
  SUPABASE_SERVICE_KEY=your_service_role_key \
  python3 upload_opg_and_build_excel.py
"""

import os, re, sys, uuid, glob
import openpyxl
from openpyxl import Workbook
from pathlib import Path

INPUT_EXCEL  = '/Users/sittminthar/Downloads/Dental-Blinding-Process-Project-master/Blinding_TT1_mapped_links.xlsx'
OUTPUT_EXCEL = '/Users/sittminthar/Downloads/Dental-Blinding-Process-Project-master/Blinding_TT1_supabase_urls.xlsx'
BUCKET = 'opg-images'
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'}


def get_creds():
    url = os.environ.get('SUPABASE_URL', '').rstrip('/')
    key = os.environ.get('SUPABASE_SERVICE_KEY') or os.environ.get('SUPABASE_KEY', '')
    if not url or not key:
        print("ERROR: set SUPABASE_URL and SUPABASE_SERVICE_KEY env vars.")
        sys.exit(1)
    return url, key


def find_image_in_folder(folder_path):
    """Return the first image file inside a folder (ignore .DS_Store, subfolders etc)."""
    if os.path.isfile(folder_path):
        # It's already a direct file
        ext = Path(folder_path).suffix.lower()
        if ext in IMAGE_EXTS:
            return folder_path
    if os.path.isdir(folder_path):
        for entry in sorted(os.listdir(folder_path)):
            full = os.path.join(folder_path, entry)
            if os.path.isfile(full) and Path(full).suffix.lower() in IMAGE_EXTS:
                return full
    return None


def parse_hyperlink(cell_val):
    """Extract path or URL from =HYPERLINK("...", "OPG") formula."""
    if not cell_val:
        return None
    s = str(cell_val).strip()
    if s in ('#VALUE!', '#REF!', ''):
        return None
    m = re.search(r'=HYPERLINK\("([^"]+)"', s, re.IGNORECASE)
    if m:
        raw = m.group(1)
        if raw.startswith('file:///'):
            return raw[7:]
        if raw.startswith('file://'):
            return raw[6:]
        return raw
    if s.startswith('http://') or s.startswith('https://'):
        return s
    return None


def upload_image(img_path, supabase_url, key):
    """Upload image to Supabase and return its signed URL (1-year expiry)."""
    import requests
    ext = Path(img_path).suffix.lstrip('.').lower() or 'jpeg'
    if ext == 'jpg':
        ext = 'jpeg'
    remote = f"opg/{uuid.uuid4().hex}.{ext}"

    with open(img_path, 'rb') as f:
        data = f.read()

    # Upload
    resp = requests.post(
        f"{supabase_url}/storage/v1/object/{BUCKET}/{remote}",
        headers={'Authorization': f'Bearer {key}', 'Content-Type': f'image/{ext}'},
        data=data
    )
    if not resp.ok:
        raise Exception(f"Upload failed {resp.status_code}: {resp.text}")

    # Get signed URL (1 year = 31536000s)
    sign_resp = requests.post(
        f"{supabase_url}/storage/v1/object/sign/{BUCKET}/{remote}",
        headers={'Authorization': f'Bearer {key}', 'Content-Type': 'application/json'},
        json={"expiresIn": 31536000}
    )
    if sign_resp.ok:
        signed = (sign_resp.json().get('signedURL') or
                  sign_resp.json().get('signedUrl') or
                  sign_resp.json().get('url', ''))
        if signed and signed.startswith('/'):
            signed = f"{supabase_url}{signed}"
        if signed:
            return signed

    # Fallback: construct public URL
    return f"{supabase_url}/storage/v1/object/public/{BUCKET}/{remote}"


def safe_age(raw):
    """Round to 1 decimal place."""
    try:
        return round(float(str(raw)), 1)
    except:
        return 0.0


def norm_sex(raw):
    s = str(raw).strip().lower() if raw else ''
    if s in ('m', 'male'):   return 'male'
    if s in ('f', 'female'): return 'female'
    return s


def main():
    supabase_url, key = get_creds()

    print(f"Reading: {INPUT_EXCEL}")
    wb_formula = openpyxl.load_workbook(INPUT_EXCEL, data_only=False)
    wb_data    = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
    ws_formula = wb_formula.active
    ws_data    = wb_data.active
    data_rows  = {i: r for i, r in enumerate(ws_data.iter_rows(min_row=1, values_only=True), 1)}

    wb_out = Workbook()
    ws_out = wb_out.active

    headers = [c.value for c in ws_formula[1]]
    ws_out.append(headers)

    total = uploaded = skipped = errors = 0

    for row_idx, formula_row in enumerate(ws_formula.iter_rows(min_row=2, values_only=True), 2):
        data_row = data_rows.get(row_idx, formula_row)
        total += 1

        out = list(data_row) if data_row else list(formula_row)

        # Fix: age → 1 decimal place
        if len(out) > 2 and out[2] is not None:
            out[2] = safe_age(out[2])

        # Fix: sex normalization
        if len(out) > 3 and out[3] is not None:
            out[3] = norm_sex(out[3])

        # OPG: resolve path from formula
        opg_url = None
        opg_formula_val = formula_row[4] if len(formula_row) > 4 else None
        resolved = parse_hyperlink(opg_formula_val)

        if resolved:
            if resolved.startswith('http'):
                opg_url = resolved
                skipped += 1
            else:
                # resolved may be a folder — find the image inside it
                img_file = find_image_in_folder(resolved)
                if img_file:
                    try:
                        name = Path(img_file).name
                        print(f"  Row {row_idx}: uploading '{name}' ...", end=' ', flush=True)
                        opg_url = upload_image(img_file, supabase_url, key)
                        uploaded += 1
                        print(f"✓")
                    except Exception as e:
                        print(f"✗ {e}")
                        errors += 1
                else:
                    print(f"  Row {row_idx}: no image found in '{resolved}'")
                    errors += 1
        else:
            # No link at all
            pass

        if len(out) > 4:
            out[4] = opg_url
        else:
            while len(out) < 5:
                out.append(None)
            out[4] = opg_url

        ws_out.append(out)

    wb_out.save(OUTPUT_EXCEL)
    wb_formula.close()
    wb_data.close()

    print()
    print("=" * 55)
    print(f"Output: {OUTPUT_EXCEL}")
    print(f"  Total rows : {total}")
    print(f"  OPG uploaded OK  : {uploaded}")
    print(f"  Already URLs     : {skipped}")
    print(f"  Errors / missing : {errors}")
    print()
    print("Import Blinding_TT1_supabase_urls.xlsx into the website.")

if __name__ == '__main__':
    main()
